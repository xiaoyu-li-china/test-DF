#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import csv
import pytest
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from rfid_inventory import (
    load_rfid_scan,
    load_wms_sku,
    load_antenna_map,
    compare_inventory,
    generate_diff_report,
    generate_markdown_summary,
    _build_excel_row,
    EXCEL_COLUMNS,
)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@pytest.fixture
def temp_csv_dir():
    """创建临时目录用于存放测试CSV文件。"""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


@pytest.fixture
def sample_rfid_csv(temp_csv_dir):
    """创建模拟的RFID扫描CSV文件。"""
    csv_path = temp_csv_dir / "rfid_scan.csv"
    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['EPC', '扫描时间', '天线号', 'RSSI'])
        writer.writerow(['EPC001', '2026-06-01 09:00:00', '1', '-50'])
        writer.writerow(['EPC002', '2026-06-01 09:00:01', '1', '-51'])
        writer.writerow(['EPC003', '2026-06-01 09:00:02', '2', '-52'])
        writer.writerow(['EPC001', '2026-06-01 09:00:03', '1', '-49'])
        writer.writerow(['EPC002', '2026-06-01 09:00:04', '1', '-48'])
        writer.writerow(['EPC004', '2026-06-01 09:00:05', '2', '-53'])
    return csv_path


@pytest.fixture
def sample_wms_csv(temp_csv_dir):
    """创建模拟的WMS账面CSV文件。"""
    csv_path = temp_csv_dir / "wms_sku.csv"
    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['EPC', 'SKU编码', 'SKU名称', '账面数量', '库位', '品类'])
        writer.writerow(['EPC001', 'SKU001', '商品A', '2', 'A-01', '电子产品'])
        writer.writerow(['EPC002', 'SKU002', '商品B', '3', 'A-02', '配件'])
        writer.writerow(['EPC003', 'SKU003', '商品C', '1', 'B-01', '办公用品'])
        writer.writerow(['EPC005', 'SKU005', '商品E', '1', 'B-02', '办公用品'])
    return csv_path


@pytest.fixture
def sample_antenna_map_csv(temp_csv_dir):
    """创建模拟的天线映射CSV文件。"""
    csv_path = temp_csv_dir / "antenna_map.csv"
    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['天线号', '货位'])
        writer.writerow(['1', 'ZONE-A'])
        writer.writerow(['2', 'ZONE-B'])
    return csv_path


class TestCsvParsing:
    """测试CSV解析功能。"""

    def test_load_rfid_scan_basic(self, sample_rfid_csv):
        """测试基本的RFID CSV解析功能。"""
        records = load_rfid_scan(sample_rfid_csv, antenna_map=None)
        
        assert len(records) == 4
        assert 'EPC001' in records
        assert 'EPC002' in records
        assert 'EPC003' in records
        assert 'EPC004' in records
        
        assert records['EPC001']['first_seen'] == '2026-06-01 09:00:00'
        assert records['EPC001']['last_seen'] == '2026-06-01 09:00:03'
        assert records['EPC001']['antenna'] == '1'

    def test_load_rfid_scan_english_columns(self, temp_csv_dir):
        """测试英文列名的RFID CSV解析。"""
        csv_path = temp_csv_dir / "rfid_english.csv"
        with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['EPC', 'ScanTime', 'Antenna', 'RSSI'])
            writer.writerow(['EPC999', '2026-06-01 10:00:00', '3', '-55'])
        
        records = load_rfid_scan(csv_path, antenna_map=None)
        assert len(records) == 1
        assert records['EPC999']['first_seen'] == '2026-06-01 10:00:00'
        assert records['EPC999']['antenna'] == '3'

    def test_load_wms_sku_basic(self, sample_wms_csv):
        """测试基本的WMS CSV解析功能。"""
        skus = load_wms_sku(sample_wms_csv)
        
        assert len(skus) == 4
        assert 'EPC001' in skus
        assert 'EPC005' in skus
        
        assert skus['EPC001']['sku_code'] == 'SKU001'
        assert skus['EPC001']['sku_name'] == '商品A'
        assert skus['EPC001']['quantity'] == 2
        assert skus['EPC001']['location'] == 'A-01'
        assert skus['EPC001']['category'] == '电子产品'

    def test_load_wms_sku_empty_quantity(self, temp_csv_dir):
        """测试空数量的处理。"""
        csv_path = temp_csv_dir / "wms_empty_qty.csv"
        with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['EPC', 'SKU编码', 'SKU名称', '账面数量', '库位', '品类'])
            writer.writerow(['EPC001', 'SKU001', '商品A', '', 'A-01', '电子产品'])
        
        skus = load_wms_sku(csv_path)
        assert skus['EPC001']['quantity'] == 1

    def test_load_antenna_map_basic(self, sample_antenna_map_csv):
        """测试天线映射表解析。"""
        antenna_map = load_antenna_map(sample_antenna_map_csv)
        
        assert len(antenna_map) == 2
        assert antenna_map['1'] == 'ZONE-A'
        assert antenna_map['2'] == 'ZONE-B'

    def test_load_antenna_map_none_path(self):
        """测试None路径的处理。"""
        antenna_map = load_antenna_map(None)
        assert antenna_map == {}


class TestEpcDeduplication:
    """测试EPC去重功能。"""

    def test_epc_deduplication_count(self, sample_rfid_csv):
        """测试同一EPC多次扫描只算一个唯一标签。"""
        records = load_rfid_scan(sample_rfid_csv, antenna_map=None)
        
        assert len(records) == 4
        assert records['EPC001']['count'] == 2
        assert records['EPC002']['count'] == 2
        assert records['EPC003']['count'] == 1

    def test_epc_deduplication_timestamps(self, sample_rfid_csv):
        """测试去重后保留正确的首次和最后扫描时间。"""
        records = load_rfid_scan(sample_rfid_csv, antenna_map=None)
        
        assert records['EPC001']['first_seen'] == '2026-06-01 09:00:00'
        assert records['EPC001']['last_seen'] == '2026-06-01 09:00:03'

    def test_surplus_count_by_unique_epc(self, sample_rfid_csv, sample_wms_csv):
        """测试盘盈数量按唯一EPC统计。"""
        rfid_records = load_rfid_scan(sample_rfid_csv, antenna_map=None)
        wms_skus = load_wms_sku(sample_wms_csv)
        result = compare_inventory(rfid_records, wms_skus)
        
        assert len(result['surplus']) == 1
        assert result['surplus'][0]['epc'] == 'EPC004'
        assert result['surplus'][0]['actual_qty'] == 1


class TestSurplusShortageClassification:
    """测试盘盈盘亏分类功能。"""

    def test_surplus_items(self, sample_rfid_csv, sample_wms_csv):
        """测试盘盈物品识别。"""
        rfid_records = load_rfid_scan(sample_rfid_csv, antenna_map=None)
        wms_skus = load_wms_sku(sample_wms_csv)
        result = compare_inventory(rfid_records, wms_skus)
        
        surplus_epcs = {item['epc'] for item in result['surplus']}
        assert surplus_epcs == {'EPC004'}

    def test_shortage_items(self, sample_rfid_csv, sample_wms_csv):
        """测试盘亏物品识别。"""
        rfid_records = load_rfid_scan(sample_rfid_csv, antenna_map=None)
        wms_skus = load_wms_sku(sample_wms_csv)
        result = compare_inventory(rfid_records, wms_skus)
        
        shortage_epcs = {item['epc'] for item in result['shortage']}
        assert shortage_epcs == {'EPC005'}

    def test_matched_items(self, sample_rfid_csv, sample_wms_csv):
        """测试匹配物品识别。"""
        rfid_records = load_rfid_scan(sample_rfid_csv, antenna_map=None)
        wms_skus = load_wms_sku(sample_wms_csv)
        result = compare_inventory(rfid_records, wms_skus)
        
        matched_epcs = {item['epc'] for item in result['match']}
        assert matched_epcs == {'EPC001', 'EPC002', 'EPC003'}

    def test_shortage_item_has_wms_data(self, sample_rfid_csv, sample_wms_csv):
        """测试盘亏物品保留WMS信息。"""
        rfid_records = load_rfid_scan(sample_rfid_csv, antenna_map=None)
        wms_skus = load_wms_sku(sample_wms_csv)
        result = compare_inventory(rfid_records, wms_skus)
        
        shortage_item = next(item for item in result['shortage'] if item['epc'] == 'EPC005')
        assert shortage_item['sku_code'] == 'SKU005'
        assert shortage_item['sku_name'] == '商品E'
        assert shortage_item['quantity'] == 1
        assert shortage_item['count'] == 0
        assert shortage_item['status'] == '盘亏'

    def test_surplus_item_has_rfid_data(self, sample_rfid_csv, sample_wms_csv):
        """测试盘盈物品保留RFID信息。"""
        rfid_records = load_rfid_scan(sample_rfid_csv, antenna_map=None)
        wms_skus = load_wms_sku(sample_wms_csv)
        result = compare_inventory(rfid_records, wms_skus)
        
        surplus_item = result['surplus'][0]
        assert surplus_item['first_seen'] == '2026-06-01 09:00:05'
        assert surplus_item['antenna'] == '2'
        assert surplus_item['status'] == '盘盈'


class TestAntennaMapping:
    """测试天线映射功能。"""

    def test_antenna_mapping_applied(self, sample_rfid_csv, sample_antenna_map_csv):
        """测试已映射天线的货位识别。"""
        antenna_map = load_antenna_map(sample_antenna_map_csv)
        records = load_rfid_scan(sample_rfid_csv, antenna_map)
        
        assert records['EPC001']['antenna_mapped'] == True
        assert records['EPC001']['mapped_location'] == 'ZONE-A'
        assert records['EPC003']['antenna_mapped'] == True
        assert records['EPC003']['mapped_location'] == 'ZONE-B'

    def test_antenna_not_mapped(self, sample_rfid_csv):
        """测试未映射天线的识别（未提供映射表）。"""
        records = load_rfid_scan(sample_rfid_csv, antenna_map={})
        
        assert records['EPC001']['antenna_mapped'] == False
        assert records['EPC001']['mapped_location'] == ''

    def test_antenna_partially_mapped(self, temp_csv_dir, sample_antenna_map_csv):
        """测试部分天线未映射时的处理。"""
        rfid_path = temp_csv_dir / "rfid_multi_antenna.csv"
        with open(rfid_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['EPC', '扫描时间', '天线号', 'RSSI'])
            writer.writerow(['EPC001', '2026-06-01 09:00:00', '1', '-50'])
            writer.writerow(['EPC001', '2026-06-01 09:00:01', '3', '-51'])
            writer.writerow(['EPC002', '2026-06-01 09:00:02', '1', '-52'])
        
        antenna_map = load_antenna_map(sample_antenna_map_csv)
        records = load_rfid_scan(rfid_path, antenna_map)
        
        assert records['EPC001']['antenna_mapped'] == False
        assert '3' in records['EPC001']['unmapped_antennas']
        assert records['EPC002']['antenna_mapped'] == True

    def test_unmapped_items_bucket(self, temp_csv_dir, sample_wms_csv, sample_antenna_map_csv):
        """测试未映射天线物品进入未映射桶。"""
        rfid_path = temp_csv_dir / "rfid_unmapped.csv"
        with open(rfid_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['EPC', '扫描时间', '天线号', 'RSSI'])
            writer.writerow(['EPC001', '2026-06-01 09:00:00', '1', '-50'])
            writer.writerow(['EPC001', '2026-06-01 09:00:01', '3', '-51'])
            writer.writerow(['EPC006', '2026-06-01 09:00:02', '3', '-52'])
        
        antenna_map = load_antenna_map(sample_antenna_map_csv)
        rfid_records = load_rfid_scan(rfid_path, antenna_map)
        wms_skus = load_wms_sku(sample_wms_csv)
        result = compare_inventory(rfid_records, wms_skus)
        
        unmapped_epcs = {item['epc'] for item in result['unmapped']}
        assert 'EPC001' in unmapped_epcs
        assert 'EPC006' in unmapped_epcs

    def test_surplus_with_mapped_location(self, sample_rfid_csv, sample_wms_csv, sample_antenna_map_csv):
        """测试盘盈物品的映射货位。"""
        antenna_map = load_antenna_map(sample_antenna_map_csv)
        rfid_records = load_rfid_scan(sample_rfid_csv, antenna_map)
        wms_skus = load_wms_sku(sample_wms_csv)
        result = compare_inventory(rfid_records, wms_skus)
        
        surplus_item = result['surplus'][0]
        assert surplus_item['mapped_location'] == 'ZONE-B'
        assert surplus_item['antenna_mapped'] == True


class TestDryRun:
    """测试dry-run模式（通过main函数参数控制）。"""

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_generate_diff_report_creates_excel(self, temp_csv_dir, sample_rfid_csv, sample_wms_csv, sample_antenna_map_csv):
        """测试正常模式下生成Excel差异报告文件。"""
        antenna_map = load_antenna_map(sample_antenna_map_csv)
        rfid_records = load_rfid_scan(sample_rfid_csv, antenna_map)
        wms_skus = load_wms_sku(sample_wms_csv)
        result = compare_inventory(rfid_records, wms_skus)
        
        output_path = temp_csv_dir / "diff_report.xlsx"
        generate_diff_report(result, output_path)
        
        assert output_path.exists()
        assert output_path.stat().st_size > 0

    def test_generate_markdown_summary_creates_file(self, temp_csv_dir, sample_rfid_csv, sample_wms_csv, sample_antenna_map_csv):
        """测试正常模式下生成Markdown摘要文件。"""
        antenna_map = load_antenna_map(sample_antenna_map_csv)
        rfid_records = load_rfid_scan(sample_rfid_csv, antenna_map)
        wms_skus = load_wms_sku(sample_wms_csv)
        result = compare_inventory(rfid_records, wms_skus)
        
        output_path = temp_csv_dir / "summary.md"
        generate_markdown_summary(result, str(sample_rfid_csv), str(sample_wms_csv), str(sample_antenna_map_csv), output_path)
        
        assert output_path.exists()
        assert output_path.stat().st_size > 0

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_excel_has_multiple_sheets(self, temp_csv_dir, sample_rfid_csv, sample_wms_csv, sample_antenna_map_csv):
        """测试Excel包含多个Sheet。"""
        antenna_map = load_antenna_map(sample_antenna_map_csv)
        rfid_records = load_rfid_scan(sample_rfid_csv, antenna_map)
        wms_skus = load_wms_sku(sample_wms_csv)
        result = compare_inventory(rfid_records, wms_skus)
        
        output_path = temp_csv_dir / "diff_report.xlsx"
        generate_diff_report(result, output_path)
        
        wb = openpyxl.load_workbook(output_path)
        sheet_names = wb.sheetnames
        
        assert '全部明细' in sheet_names
        assert '差异项' in sheet_names
        assert '盘盈' in sheet_names
        assert '盘亏' in sheet_names
        wb.close()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_excel_all_sheet_contains_all_statuses(self, temp_csv_dir, sample_rfid_csv, sample_wms_csv, sample_antenna_map_csv):
        """测试Excel全部明细Sheet包含所有状态。"""
        antenna_map = load_antenna_map(sample_antenna_map_csv)
        rfid_records = load_rfid_scan(sample_rfid_csv, antenna_map)
        wms_skus = load_wms_sku(sample_wms_csv)
        result = compare_inventory(rfid_records, wms_skus)
        
        output_path = temp_csv_dir / "diff_report.xlsx"
        generate_diff_report(result, output_path)
        
        wb = openpyxl.load_workbook(output_path)
        ws = wb['全部明细']
        
        status_col = None
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col_idx).value == '状态':
                status_col = col_idx
                break
        
        assert status_col is not None
        statuses = set()
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=status_col).value
            if val:
                statuses.add(val)
        
        assert '相符' in statuses
        assert '盘盈' in statuses
        assert '盘亏' in statuses
        wb.close()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_excel_surplus_sheet_only_surplus(self, temp_csv_dir, sample_rfid_csv, sample_wms_csv, sample_antenna_map_csv):
        """测试盘盈Sheet只包含盘盈物品。"""
        antenna_map = load_antenna_map(sample_antenna_map_csv)
        rfid_records = load_rfid_scan(sample_rfid_csv, antenna_map)
        wms_skus = load_wms_sku(sample_wms_csv)
        result = compare_inventory(rfid_records, wms_skus)
        
        output_path = temp_csv_dir / "diff_report.xlsx"
        generate_diff_report(result, output_path)
        
        wb = openpyxl.load_workbook(output_path)
        ws = wb['盘盈']
        
        status_col = None
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col_idx).value == '状态':
                status_col = col_idx
                break
        
        for row_idx in range(2, ws.max_row + 1):
            assert ws.cell(row=row_idx, column=status_col).value == '盘盈'
        wb.close()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_excel_shortage_sheet_only_shortage(self, temp_csv_dir, sample_rfid_csv, sample_wms_csv, sample_antenna_map_csv):
        """测试盘亏Sheet只包含盘亏物品。"""
        antenna_map = load_antenna_map(sample_antenna_map_csv)
        rfid_records = load_rfid_scan(sample_rfid_csv, antenna_map)
        wms_skus = load_wms_sku(sample_wms_csv)
        result = compare_inventory(rfid_records, wms_skus)
        
        output_path = temp_csv_dir / "diff_report.xlsx"
        generate_diff_report(result, output_path)
        
        wb = openpyxl.load_workbook(output_path)
        ws = wb['盘亏']
        
        status_col = None
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col_idx).value == '状态':
                status_col = col_idx
                break
        
        for row_idx in range(2, ws.max_row + 1):
            assert ws.cell(row=row_idx, column=status_col).value == '盘亏'
        wb.close()

    @pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
    def test_excel_status_cell_has_color(self, temp_csv_dir, sample_rfid_csv, sample_wms_csv, sample_antenna_map_csv):
        """测试状态列有条件格式化颜色。"""
        antenna_map = load_antenna_map(sample_antenna_map_csv)
        rfid_records = load_rfid_scan(sample_rfid_csv, antenna_map)
        wms_skus = load_wms_sku(sample_wms_csv)
        result = compare_inventory(rfid_records, wms_skus)
        
        output_path = temp_csv_dir / "diff_report.xlsx"
        generate_diff_report(result, output_path)
        
        wb = openpyxl.load_workbook(output_path)
        ws = wb['全部明细']
        
        status_col = None
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col_idx).value == '状态':
                status_col = col_idx
                break
        
        colored_rows = 0
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=status_col)
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb and cell.fill.start_color.rgb != '00000000':
                colored_rows += 1
        
        assert colored_rows > 0
        wb.close()

    def test_markdown_summary_contains_sections(self, temp_csv_dir, sample_rfid_csv, sample_wms_csv, sample_antenna_map_csv):
        """测试Markdown摘要包含所有章节。"""
        antenna_map = load_antenna_map(sample_antenna_map_csv)
        rfid_records = load_rfid_scan(sample_rfid_csv, antenna_map)
        wms_skus = load_wms_sku(sample_wms_csv)
        result = compare_inventory(rfid_records, wms_skus)
        
        output_path = temp_csv_dir / "summary.md"
        generate_markdown_summary(result, str(sample_rfid_csv), str(sample_wms_csv), str(sample_antenna_map_csv), output_path)
        
        with open(output_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        assert '## 一、盘点概览' in content
        assert '## 二、差异明细' in content
        assert '### 2.1 盘盈物品' in content
        assert '### 2.2 盘亏物品' in content
        assert '## 三、分类统计' in content
        assert '盘亏' in content

    def test_empty_input_handling(self, temp_csv_dir):
        """测试空输入文件的处理。"""
        rfid_path = temp_csv_dir / "empty_rfid.csv"
        wms_path = temp_csv_dir / "empty_wms.csv"
        
        with open(rfid_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['EPC', '扫描时间', '天线号', 'RSSI'])
        
        with open(wms_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['EPC', 'SKU编码', 'SKU名称', '账面数量', '库位', '品类'])
        
        rfid_records = load_rfid_scan(rfid_path, antenna_map=None)
        wms_skus = load_wms_sku(wms_path)
        result = compare_inventory(rfid_records, wms_skus)
        
        assert len(result['match']) == 0
        assert len(result['surplus']) == 0
        assert len(result['shortage']) == 0


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
