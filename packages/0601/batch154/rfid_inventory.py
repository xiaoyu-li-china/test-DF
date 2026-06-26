#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import csv
import sys
import json
import argparse
import tempfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

try:
    import paramiko
    HAS_PARAMIKO = True
except ImportError:
    HAS_PARAMIKO = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def load_sftp_config(config_path):
    """读取SFTP配置文件。"""
    if not config_path:
        return None
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def fetch_wms_from_sftp(sftp_config, local_path=None):
    """从WMS SFTP服务器拉取nightly账面清单。"""
    if not HAS_PARAMIKO:
        raise ImportError("需要安装 paramiko: pip install paramiko")
    
    host = sftp_config['host']
    port = sftp_config.get('port', 22)
    username = sftp_config['username']
    password = sftp_config.get('password')
    private_key = sftp_config.get('private_key')
    remote_path = sftp_config['remote_path']
    
    if local_path is None:
        local_path = tempfile.NamedTemporaryFile(suffix='.csv', delete=False).name
    
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    
    try:
        if private_key:
            key = paramiko.RSAKey.from_private_key_file(private_key)
            ssh.connect(host, port=port, username=username, pkey=key)
        else:
            ssh.connect(host, port=port, username=username, password=password)
        
        sftp = ssh.open_sftp()
        sftp.get(remote_path, local_path)
        sftp.close()
        ssh.close()
        
        return local_path
    except Exception as e:
        raise RuntimeError(f"SFTP拉取失败: {str(e)}")


def send_slack_notification(webhook_url, result, top_n=10):
    """推送盘亏Top10到Slack。"""
    if not webhook_url:
        return False
    
    if not HAS_REQUESTS:
        print("⚠️  缺少 requests 库，无法推送Slack通知")
        return False
    
    shortage_items = result['shortage']
    shortage_items_sorted = sorted(shortage_items, key=lambda x: x['quantity'], reverse=True)[:top_n]
    
    total_shortage = len(shortage_items)
    total_qty = sum(x['quantity'] for x in shortage_items)
    
    blocks = [
        {
            "type": "header",
            "text": {
                "type": "plain_text",
                "text": "📦 RFID月度盘点 - 盘亏预警"
            }
        },
        {
            "type": "section",
            "fields": [
                {
                    "type": "mrkdwn",
                    "text": f"*盘亏总数:*\n{total_shortage} 个SKU"
                },
                {
                    "type": "mrkdwn",
                    "text": f"*盘亏总数量:*\n{total_qty} 件"
                }
            ]
        },
        {
            "type": "divider"
        },
        {
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": f"*盘亏 Top {min(top_n, len(shortage_items_sorted))} 明细:*"
            }
        }
    ]
    
    if shortage_items_sorted:
        for item in shortage_items_sorted:
            blocks.append({
                "type": "section",
                "fields": [
                    {
                        "type": "mrkdwn",
                        "text": f"*SKU:* `{item['sku_code']}`\n*名称:* {item['sku_name']}"
                    },
                    {
                        "type": "mrkdwn",
                        "text": f"*库位:* {item['location']}\n*数量:* `{item['quantity']}` 件"
                    }
                ]
            })
    else:
        blocks.append({
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": "✅ 无盘亏物品！"
            }
        })
    
    blocks.extend([
        {
            "type": "divider"
        },
        {
            "type": "context",
            "elements": [
                {
                    "type": "mrkdwn",
                    "text": f"盘点时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                }
            ]
        }
    ])
    
    payload = {
        "text": f"盘亏预警: {total_shortage} 个SKU, {total_qty} 件",
        "blocks": blocks
    }
    
    try:
        response = requests.post(webhook_url, json=payload, timeout=10)
        return response.status_code == 200
    except Exception as e:
        print(f"⚠️  Slack推送失败: {str(e)}")
        return False


def load_antenna_map(csv_path):
    """读取天线号到货位的映射表。"""
    antenna_map = {}
    if not csv_path:
        return antenna_map
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            antenna = str(row.get('天线号', row.get('Antenna', '')).strip())
            location = row.get('货位', row.get('Location', '')).strip()
            if antenna:
                antenna_map[antenna] = location
    return antenna_map


def load_rfid_scan(csv_path, antenna_map=None):
    """读取手持枪导出的CSV文件，返回EPC到扫描记录的映射。"""
    records = {}
    antenna_map = antenna_map or {}
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            epc = row.get('EPC', '').strip()
            if not epc:
                continue
            scan_time = row.get('扫描时间', row.get('ScanTime', '')).strip()
            antenna = str(row.get('天线号', row.get('Antenna', '')).strip())
            rssi = row.get('RSSI', '').strip()
            is_antenna_mapped = antenna in antenna_map
            mapped_location = antenna_map.get(antenna, '')

            if epc not in records:
                records[epc] = {
                    'epc': epc,
                    'first_seen': scan_time,
                    'last_seen': scan_time,
                    'antenna': antenna,
                    'rssi': rssi,
                    'count': 1,
                    'mapped_location': mapped_location,
                    'antenna_mapped': is_antenna_mapped,
                    'all_antennas': {antenna},
                    'unmapped_antennas': set() if is_antenna_mapped else {antenna}
                }
            else:
                records[epc]['last_seen'] = scan_time
                records[epc]['count'] += 1
                records[epc]['all_antennas'].add(antenna)
                if not is_antenna_mapped:
                    records[epc]['unmapped_antennas'].add(antenna)
                    records[epc]['antenna_mapped'] = False
                if not records[epc]['mapped_location'] and mapped_location:
                    records[epc]['mapped_location'] = mapped_location
                    records[epc]['antenna'] = antenna
    return records


def load_wms_sku(csv_path):
    """读取WMS导出的账面SKU清单，返回EPC到SKU信息的映射。"""
    skus = {}
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            epc = row.get('EPC', '').strip()
            if not epc:
                continue
            skus[epc] = {
                'epc': epc,
                'sku_code': row.get('SKU编码', row.get('SKU', '')).strip(),
                'sku_name': row.get('SKU名称', row.get('Name', '')).strip(),
                'quantity': int(row.get('账面数量', row.get('Quantity', '1')).strip() or 1),
                'location': row.get('库位', row.get('Location', '')).strip(),
                'category': row.get('品类', row.get('Category', '')).strip()
            }
    return skus


def compare_inventory(rfid_records, wms_skus):
    """比对RFID扫描数据和WMS账面数据，生成差异结果。"""
    result = {
        'match': [],
        'surplus': [],
        'shortage': [],
        'unscanned': [],
        'unmapped': []
    }

    rfid_epcs = set(rfid_records.keys())
    wms_epcs = set(wms_skus.keys())

    matched_epcs = rfid_epcs & wms_epcs
    surplus_epcs = rfid_epcs - wms_epcs
    shortage_epcs = wms_epcs - rfid_epcs

    for epc in matched_epcs:
        item = {
            **wms_skus[epc],
            **rfid_records[epc],
            'status': '相符',
            'rfid_count': rfid_records[epc]['count'],
            'actual_qty': 1
        }
        if not rfid_records[epc]['antenna_mapped']:
            result['unmapped'].append(item)
        result['match'].append(item)

    for epc in surplus_epcs:
        rfid_data = rfid_records[epc]
        item = {
            'epc': epc,
            'sku_code': '',
            'sku_name': '',
            'quantity': 0,
            'location': rfid_data['mapped_location'],
            'category': '',
            **rfid_data,
            'status': '盘盈',
            'rfid_count': rfid_data['count'],
            'actual_qty': 1
        }
        if not rfid_data['antenna_mapped']:
            result['unmapped'].append(item)
        result['surplus'].append(item)

    for epc in shortage_epcs:
        wms_data = wms_skus[epc]
        item = {
            **wms_data,
            'first_seen': '',
            'last_seen': '',
            'antenna': '',
            'rssi': '',
            'count': 0,
            'mapped_location': '',
            'antenna_mapped': True,
            'status': '盘亏',
            'rfid_count': 0,
            'actual_qty': 0
        }
        result['shortage'].append(item)

    for item in result['match']:
        if item['count'] < item['quantity']:
            item['status'] = '数量不足'
            result['unscanned'].append(item)

    return result


def _build_excel_row(item):
    """将比对结果项转换为Excel行字典。"""
    all_ants = ','.join(sorted(item.get('all_antennas', {item.get('antenna', '')})))
    unmapped_ants = ','.join(sorted(item.get('unmapped_antennas', set())))
    return {
        'EPC': item['epc'],
        'SKU编码': item['sku_code'],
        'SKU名称': item['sku_name'],
        '账面数量': item['quantity'],
        '实际数量': item['actual_qty'],
        '扫描次数': item['count'],
        '库位(WMS)': item['location'],
        '映射货位(天线)': item.get('mapped_location', ''),
        '品类': item['category'],
        '首次扫描时间': item['first_seen'],
        '最后扫描时间': item['last_seen'],
        '所有天线': all_ants,
        '未映射天线': unmapped_ants,
        'RSSI': item['rssi'],
        '状态': item['status'],
        '天线映射': '已映射' if item.get('antenna_mapped', True) else '未映射'
    }


EXCEL_COLUMNS = [
    'EPC', 'SKU编码', 'SKU名称', '账面数量', '实际数量', '扫描次数',
    '库位(WMS)', '映射货位(天线)', '品类', '首次扫描时间', '最后扫描时间',
    '所有天线', '未映射天线', 'RSSI', '状态', '天线映射'
]

STATUS_FILLS = {
    '相符': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    '盘盈': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    '盘亏': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    '数量不足': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
}

STATUS_FONTS = {
    '相符': Font(color='006100'),
    '盘盈': Font(color='9C5700'),
    '盘亏': Font(color='9C0006'),
    '数量不足': Font(color='9C0006'),
}


def generate_diff_report(result, output_path):
    """生成差异报告Excel文件，含多Sheet和条件格式化。"""
    if not HAS_OPENPYXL:
        raise ImportError("需要安装 openpyxl: pip install openpyxl")

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def _write_sheet(ws, items, col_widths=None):
        for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, item in enumerate(items, 2):
            row_data = _build_excel_row(item)
            for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
                val = row_data.get(col_name, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
                status = row_data.get('状态', '')
                if col_name == '状态' and status in STATUS_FILLS:
                    cell.fill = STATUS_FILLS[status]
                    cell.font = STATUS_FONTS[status]
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        if col_widths:
            for col_idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'

    all_items = result['match'] + result['surplus'] + result['shortage']
    all_items.sort(key=lambda x: (x['status'], x['sku_code'] or '', x['epc']))

    ws_all = wb.active
    ws_all.title = '全部明细'
    _write_sheet(ws_all, all_items, col_widths=[
        28, 12, 22, 10, 10, 10, 12, 14, 12, 20, 20, 10, 10, 8, 10, 10
    ])

    diff_items = [i for i in all_items if i['status'] != '相符']
    if diff_items:
        ws_diff = wb.create_sheet('差异项')
        _write_sheet(ws_diff, diff_items, col_widths=[
            28, 12, 22, 10, 10, 10, 12, 14, 12, 20, 20, 10, 10, 8, 10, 10
        ])

    if result['surplus']:
        ws_surplus = wb.create_sheet('盘盈')
        _write_sheet(ws_surplus, result['surplus'], col_widths=[
            28, 12, 22, 10, 10, 10, 12, 14, 12, 20, 20, 10, 10, 8, 10, 10
        ])

    if result['shortage']:
        ws_shortage = wb.create_sheet('盘亏')
        _write_sheet(ws_shortage, result['shortage'], col_widths=[
            28, 12, 22, 10, 10, 10, 12, 14, 12, 20, 20, 10, 10, 8, 10, 10
        ])

    if result['unscanned']:
        ws_unscanned = wb.create_sheet('数量不符')
        _write_sheet(ws_unscanned, result['unscanned'], col_widths=[
            28, 12, 22, 10, 10, 10, 12, 14, 12, 20, 20, 10, 10, 8, 10, 10
        ])

    if result['unmapped']:
        ws_unmapped = wb.create_sheet('未映射天线')
        _write_sheet(ws_unmapped, result['unmapped'], col_widths=[
            28, 12, 22, 10, 10, 10, 12, 14, 12, 20, 20, 10, 10, 8, 10, 10
        ])

    wb.save(output_path)


def generate_markdown_summary(result, rfid_path, wms_path, antenna_map_path, output_path):
    """生成给主管的Markdown摘要报告。"""
    total_wms = len(result['match']) + len(result['shortage'])
    total_rfid = len(result['match']) + len(result['surplus'])
    match_count = len([x for x in result['match'] if x['status'] == '相符'])
    surplus_count = len(result['surplus'])
    shortage_count = len(result['shortage'])
    unscanned_count = len(result['unscanned'])
    unmapped_count = len(result['unmapped'])

    if total_wms > 0:
        accuracy = (match_count / total_wms) * 100
    else:
        accuracy = 0.0

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    category_stats = defaultdict(lambda: {'total': 0, 'match': 0, 'shortage': 0, 'surplus': 0})
    unmapped_stats = defaultdict(lambda: {'match': 0, 'surplus': 0})

    for item in result['match']:
        if not item['antenna_mapped']:
            for ant in item['unmapped_antennas']:
                unmapped_stats[ant]['match'] += 1
            cat = '未映射'
        else:
            cat = item['category'] or '未分类'
        category_stats[cat]['total'] += 1
        if item['status'] == '相符':
            category_stats[cat]['match'] += 1

    for item in result['shortage']:
        cat = item['category'] or '未分类'
        category_stats[cat]['total'] += 1
        category_stats[cat]['shortage'] += 1

    for item in result['surplus']:
        if not item['antenna_mapped']:
            for ant in item['unmapped_antennas']:
                unmapped_stats[ant]['surplus'] += 1
            cat = '未映射'
        else:
            cat = item.get('category', '') or '未分类'
        category_stats[cat]['surplus'] += 1

    md = f"""# RFID月度盘点报告

**生成时间**：{now}
**盘点周期**：{datetime.now().strftime('%Y年%m月')}

---

## 一、盘点概览

| 指标 | 数量 | 说明 |
|------|------|------|
| WMS账面总数 | {total_wms} | 系统记录的SKU总数 |
| RFID实际扫描 | {total_rfid} | 手持枪扫描到的唯一标签数 |
| 账实相符 | {match_count} | 账面与实际一致 |
| 盘盈 | {surplus_count} | 有实物无账面（按唯一EPC计） |
| 盘亏 | {shortage_count} | 有账面无实物 |
| 数量不符 | {unscanned_count} | 扫描次数少于账面数量 |
| 天线未映射 | {unmapped_count} | 天线号无对应货位配置 |
| **盘点准确率** | **{accuracy:.2f}%** | 相符数 / 账面总数 |

---

## 二、差异明细

### 2.1 盘盈物品（{surplus_count} 项）
*有实物但WMS系统无账面记录，按唯一EPC统计，需核实是否为漏登或错放*

"""

    if result['surplus']:
        md += "| EPC | 首次扫描时间 | 天线号 | 映射货位 | 实际数量 | 扫描次数 |\n|-----|-------------|--------|----------|----------|----------|\n"
        for item in sorted(result['surplus'], key=lambda x: x['first_seen']):
            mapped_loc = item['mapped_location'] if item['antenna_mapped'] else '**未映射**'
            md += f"| {item['epc']} | {item['first_seen']} | {item['antenna']} | {mapped_loc} | {item['actual_qty']} | {item['count']} |\n"
    else:
        md += "> ✅ 无盘盈物品\n"

    md += f"""

### 2.2 盘亏物品（{shortage_count} 项）
*WMS有账面记录但实际未扫描到，需核实是否遗失或错放*

"""

    if result['shortage']:
        md += "| EPC | SKU编码 | SKU名称 | 账面数量 | 库位 | 品类 |\n|-----|---------|---------|----------|------|------|\n"
        for item in sorted(result['shortage'], key=lambda x: (x['category'] or '', x['sku_code'] or '')):
            md += f"| {item['epc']} | {item['sku_code']} | {item['sku_name']} | {item['quantity']} | {item['location']} | {item['category']} |\n"
    else:
        md += "> ✅ 无盘亏物品\n"

    md += f"""

### 2.3 数量不符物品（{unscanned_count} 项）
*扫描到但次数少于账面数量，可能存在部分遗失*

"""

    if result['unscanned']:
        md += "| EPC | SKU编码 | SKU名称 | 账面数量 | 扫描次数 | 差异 | 库位 |\n|-----|---------|---------|----------|----------|------|------|\n"
        for item in sorted(result['unscanned'], key=lambda x: x['quantity'] - x['count'], reverse=True):
            diff = item['quantity'] - item['count']
            md += f"| {item['epc']} | {item['sku_code']} | {item['sku_name']} | {item['quantity']} | {item['count']} | **-{diff}** | {item['location']} |\n"
    else:
        md += "> ✅ 无数量不符物品\n"

    md += f"""

### 2.4 未映射天线物品（{unmapped_count} 项）
*天线号未在映射表中配置，无法确定货位，请补充映射配置*

"""

    if result['unmapped']:
        md += "| 天线号 | 未映射天线 | EPC | 状态 | 首次扫描时间 |\n|--------|-----------|-----|------|-------------|\n"
        for item in sorted(result['unmapped'], key=lambda x: (','.join(sorted(item['unmapped_antennas'])), x['epc'])):
            unmapped_ants = ','.join(sorted(item['unmapped_antennas']))
            all_ants = ','.join(sorted(item['all_antennas']))
            md += f"| {all_ants} | **{unmapped_ants}** | {item['epc']} | {item['status']} | {item['first_seen']} |\n"

        md += "\n**未映射天线统计：**\n\n"
        md += "| 天线号 | 相符数 | 盘盈数 | 合计 |\n|--------|--------|--------|------|\n"
        for ant, stats in sorted(unmapped_stats.items()):
            total = stats['match'] + stats['surplus']
            md += f"| **{ant}** | {stats['match']} | {stats['surplus']} | {total} |\n"
    else:
        md += "> ✅ 所有天线号均已配置映射\n"

    md += """

---

## 三、分类统计

| 品类 | 账面总数 | 相符 | 盘亏 | 盘盈 | 相符率 |
|------|---------|------|------|------|--------|
"""

    for cat, stats in sorted(category_stats.items()):
        rate = (stats['match'] / stats['total'] * 100) if stats['total'] > 0 else 0
        md += f"| {cat} | {stats['total']} | {stats['match']} | {stats['shortage']} | {stats['surplus']} | {rate:.2f}% |\n"

    md += f"""

---

## 四、操作说明

- **手持枪数据文件**：`{Path(rfid_path).name}`
- **WMS账面数据文件**：`{Path(wms_path).name}`
"""

    if antenna_map_path:
        md += f"- **天线映射表**：`{Path(antenna_map_path).name}`\n"
    else:
        md += "- **天线映射表**：`未配置`（建议配置以获得更准确的货位分析）\n"

    md += f"""- **详细差异报告**：`{Path(output_path).name.replace('.md', '_差异报告.csv')}`

---

## 五、后续建议

1. **盘盈物品**：建议逐一核实EPC对应的实物，确认是否为新购入未入账、错放区域或标签重复
2. **盘亏物品**：建议优先检查高频出入库区域，核实是否已出库未销账，必要时启动追查流程
3. **数量不符**：建议检查是否存在标签损坏、遮挡，或重新盘点对应库位
4. **未映射天线**：请尽快补充天线号到货位的映射配置，确保所有扫描数据可准确定位
5. **准确率{accuracy:.2f}%**：{'✅ 达到预期目标' if accuracy >= 95 else '⚠️ 建议分析原因，优化盘点流程'}

---

*报告由RFID盘点系统自动生成*
"""

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(md)


def main():
    parser = argparse.ArgumentParser(description='RFID月度盘点比对工具')
    parser.add_argument('--rfid', required=True, help='手持枪导出的CSV文件路径')
    parser.add_argument('--wms', help='WMS导出的账面SKU CSV文件路径（不指定则从SFTP拉取）')
    parser.add_argument('--sftp-config', help='WMS SFTP配置文件路径（JSON格式）')
    parser.add_argument('--antenna-map', help='天线号到货位映射表CSV文件路径')
    parser.add_argument('--slack-webhook', help='Slack Webhook URL，用于推送盘亏Top10')
    parser.add_argument('--output-dir', default='./output', help='输出目录，默认为 ./output')
    parser.add_argument('--dry-run', action='store_true', help='预览模式，只比对不生成报告文件')

    args = parser.parse_args()

    rfid_path = Path(args.rfid)
    wms_path = Path(args.wms) if args.wms else None
    sftp_config_path = Path(args.sftp_config) if args.sftp_config else None
    antenna_map_path = Path(args.antenna_map) if args.antenna_map else None
    slack_webhook = args.slack_webhook
    output_dir = Path(args.output_dir)
    dry_run = args.dry_run

    if not rfid_path.exists():
        print(f"❌ 错误：RFID扫描文件不存在: {rfid_path}")
        sys.exit(1)
    if not wms_path and not sftp_config_path:
        print(f"❌ 错误：必须指定 --wms 或 --sftp-config")
        sys.exit(1)
    if wms_path and not wms_path.exists():
        print(f"❌ 错误：WMS账面文件不存在: {wms_path}")
        sys.exit(1)
    if sftp_config_path and not sftp_config_path.exists():
        print(f"❌ 错误：SFTP配置文件不存在: {sftp_config_path}")
        sys.exit(1)
    if antenna_map_path and not antenna_map_path.exists():
        print(f"❌ 错误：天线映射表不存在: {antenna_map_path}")
        sys.exit(1)

    if dry_run:
        print("⚠️  预览模式（--dry-run）：只比对不生成报告文件")
    print("📊 RFID月度盘点比对工具")
    print("=" * 50)

    temp_wms_path = None
    if sftp_config_path:
        print(f"[1/6] 从SFTP拉取WMS账面数据: {sftp_config_path.name}")
        try:
            sftp_config = load_sftp_config(str(sftp_config_path))
            temp_wms_path = fetch_wms_from_sftp(sftp_config)
            wms_path = Path(temp_wms_path)
            print(f"      成功拉取到本地: {temp_wms_path}")
        except Exception as e:
            print(f"❌ SFTP拉取失败: {str(e)}")
            sys.exit(1)
        step_offset = 0
    else:
        step_offset = -1

    print(f"[{2+step_offset}/6] 读取天线映射表" + (f": {antenna_map_path.name}" if antenna_map_path else "（未配置）"))
    antenna_map = load_antenna_map(str(antenna_map_path) if antenna_map_path else None)
    if antenna_map:
        print(f"      已加载 {len(antenna_map)} 条天线映射规则")
    else:
        print(f"      提示：未配置天线映射，建议添加 --antenna-map 参数")

    print(f"[{3+step_offset}/6] 读取RFID扫描数据: {rfid_path.name}")
    rfid_records = load_rfid_scan(rfid_path, antenna_map)
    print(f"      共读取 {len(rfid_records)} 个唯一EPC标签")

    print(f"[{4+step_offset}/6] 读取WMS账面数据: {wms_path.name}")
    wms_skus = load_wms_sku(wms_path)
    print(f"      共读取 {len(wms_skus)} 条账面SKU记录")

    print(f"[{5+step_offset}/6] 比对库存数据...")
    result = compare_inventory(rfid_records, wms_skus)

    match_count = len([x for x in result['match'] if x['status'] == '相符'])
    print(f"      账实相符: {match_count}")
    print(f"      盘盈: {len(result['surplus'])}（按唯一EPC统计）")
    print(f"      盘亏: {len(result['shortage'])}")
    print(f"      数量不符: {len(result['unscanned'])}")
    print(f"      天线未映射: {len(result['unmapped'])}")

    total = len(result['match']) + len(result['shortage'])
    accuracy = (match_count / total * 100) if total > 0 else 0
    print(f"      盘点准确率: {accuracy:.2f}%")

    if slack_webhook:
        print(f"[{6+step_offset}/6] 推送Slack通知...")
        if send_slack_notification(slack_webhook, result, top_n=10):
            print(f"      Slack通知已发送")
        else:
            print(f"      Slack通知发送失败")
    else:
        print(f"[{6+step_offset}/6] Slack通知（未配置）")

    if not dry_run:
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        diff_report_path = output_dir / f'盘点差异报告_{timestamp}.xlsx'
        summary_path = output_dir / f'月度盘点摘要_{timestamp}.md'

        print(f"      生成差异报告: {diff_report_path.name}")
        generate_diff_report(result, diff_report_path)

        print(f"      生成摘要报告: {summary_path.name}")
        generate_markdown_summary(result, str(rfid_path), str(wms_path), str(antenna_map_path) if antenna_map_path else None, summary_path)

    print("=" * 50)
    if dry_run:
        print("✅ 预览完成！（--dry-run 模式未生成文件）")
    else:
        print("✅ 盘点完成！请查看输出目录中的报告文件。")

    total_issues = len(result['surplus']) + len(result['shortage']) + len(result['unscanned']) + len(result['unmapped'])
    if total_issues > 0:
        print(f"⚠️  发现 {total_issues} 项问题，请及时处理。")

    if temp_wms_path and not dry_run:
        saved_path = output_dir / f'wms_snapshot_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        Path(temp_wms_path).rename(saved_path)
        print(f"💾 WMS快照已保存: {saved_path}")


if __name__ == '__main__':
    main()
